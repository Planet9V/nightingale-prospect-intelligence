#!/usr/bin/env python3
"""
Output Generator for PSI System
Creates Excel and CSV outputs with proper formatting
"""

import pandas as pd
from datetime import datetime
from pathlib import Path
import logging
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class OutputGenerator:
    """Generates formatted output files"""
    
    def __init__(self):
        self.logger = logging.getLogger('OutputGenerator')
        self.version = "v1.0"
    
    def create_excel(self, personas: List[Dict], output_path: Path, batch_id: str):
        """Create formatted Excel output"""
        try:
            # Convert to DataFrame
            df = pd.DataFrame(personas)
            
            # Reorder columns
            column_order = [
                'sf_id', 'company_name', 'industry', 'sub_industry',
                'annual_revenue', 'employee_count', 'headquarters',
                'full_name', 'title', 'department', 'email', 'phone', 'linkedin',
                'role_type', 'authority_level', 'budget_authority', 'veto_power',
                'tenure', 'previous_role', 'education',
                'm_a_score', 'm_a_rationale', 'ransomware_score', 'ransomware_rationale',
                'overall_priority', 'key_priorities', 'pain_points', 'engagement_approach',
                'extraction_confidence', 'source_file'
            ]
            
            # Add missing columns with empty values
            for col in column_order:
                if col not in df.columns:
                    df[col] = ''
            
            # Reorder to match column_order
            df = df[column_order]
            
            # Add metadata columns
            df['extraction_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            df['batch_id'] = batch_id
            df['psi_version'] = self.version
            
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Main data sheet
                df.to_excel(writer, sheet_name='Stakeholder Intelligence', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Stakeholder Intelligence']
                
                # Apply formatting
                self.format_excel_sheet(worksheet, df)
                
                # Create summary sheet
                summary_df = self.create_summary(df)
                summary_df.to_excel(writer, sheet_name='Summary', index=True)
                
                # Format summary sheet
                summary_ws = writer.sheets['Summary']
                self.format_summary_sheet(summary_ws)
                
                # Create campaign analysis sheet
                campaign_df = self.create_campaign_analysis(df)
                campaign_df.to_excel(writer, sheet_name='Campaign Analysis', index=False)
            
            self.logger.info(f"Excel file created: {output_path}")
            
        except Exception as e:
            self.logger.error(f"Failed to create Excel: {str(e)}")
            raise
    
    def create_csv(self, personas: List[Dict], output_path: Path):
        """Create CSV backup"""
        try:
            df = pd.DataFrame(personas)
            
            # Add metadata
            df['extraction_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            df['psi_version'] = self.version
            
            # Save to CSV
            df.to_csv(output_path, index=False, encoding='utf-8')
            
            self.logger.info(f"CSV file created: {output_path}")
            
        except Exception as e:
            self.logger.error(f"Failed to create CSV: {str(e)}")
            raise
    
    def format_excel_sheet(self, worksheet, df):
        """Apply formatting to Excel worksheet"""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Freeze panes (keep header and first 3 columns visible)
        worksheet.freeze_panes = 'D2'
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Conditional formatting for scores
        # High scores (90-100) in green
        high_score_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        # Medium scores (60-89) in yellow
        med_score_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        # Low scores (0-59) in red
        low_score_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Apply conditional formatting to score columns
        score_columns = ['m_a_score', 'ransomware_score']
        for col_name in score_columns:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    try:
                        score = int(cell.value)
                        if score >= 90:
                            cell.fill = high_score_fill
                        elif score >= 60:
                            cell.fill = med_score_fill
                        else:
                            cell.fill = low_score_fill
                    except:
                        pass
    
    def format_summary_sheet(self, worksheet):
        """Format the summary sheet"""
        # Header formatting
        header_font = Font(bold=True)
        
        # Apply formatting
        for row in worksheet.iter_rows():
            row[0].font = header_font
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create summary statistics"""
        summary_data = {
            'Total Personas': len(df),
            'Total Companies': df['company_name'].nunique(),
            'Industries Covered': df['industry'].nunique(),
            
            # Role distribution
            'C-Suite Executives': len(df[df['role_type'].str.contains('CEO|CFO|CIO|CTO|CISO|COO|C-Suite', na=False)]),
            'VPs': len(df[df['role_type'].str.contains('VP|EVP|SVP', na=False)]),
            'Directors': len(df[df['role_type'] == 'Director']),
            'Managers': len(df[df['role_type'] == 'Manager']),
            
            # Authority distribution
            'Ultimate Decision Makers': len(df[df['authority_level'] == 'Ultimate Decision Maker']),
            'Executive Decision Makers': len(df[df['authority_level'] == 'Executive Decision Maker']),
            'Budget Authority': len(df[df['budget_authority'] == 'Yes']),
            
            # Campaign relevance
            'High M&A Priority (90+)': len(df[df['m_a_score'] >= 90]),
            'High Ransomware Priority (90+)': len(df[df['ransomware_score'] >= 90]),
            
            # Data completeness
            'With Email': len(df[df['email'].notna() & (df['email'] != '')]),
            'With Phone': len(df[df['phone'].notna() & (df['phone'] != '')]),
            'With LinkedIn': len(df[df['linkedin'].notna() & (df['linkedin'] != '')]),
            
            # Extraction metrics
            'Average Extraction Confidence': df['extraction_confidence'].mean() if 'extraction_confidence' in df.columns else 0,
            'Files Processed': df['source_file'].nunique() if 'source_file' in df.columns else 0
        }
        
        return pd.DataFrame(summary_data.items(), columns=['Metric', 'Value'])
    
    def create_campaign_analysis(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create campaign-specific analysis"""
        # Top M&A targets
        m_a_top = df.nlargest(20, 'm_a_score')[
            ['company_name', 'full_name', 'title', 'm_a_score', 'm_a_rationale']
        ].copy()
        m_a_top['Campaign'] = 'M&A Due Diligence'
        
        # Top Ransomware targets
        ransomware_top = df.nlargest(20, 'ransomware_score')[
            ['company_name', 'full_name', 'title', 'ransomware_score', 'ransomware_rationale']
        ].copy()
        ransomware_top['Campaign'] = 'Ransomware'
        
        # Rename score columns for consistency
        m_a_top = m_a_top.rename(columns={'m_a_score': 'Score', 'm_a_rationale': 'Rationale'})
        ransomware_top = ransomware_top.rename(columns={'ransomware_score': 'Score', 'ransomware_rationale': 'Rationale'})
        
        # Combine
        campaign_analysis = pd.concat([m_a_top, ransomware_top], ignore_index=True)
        
        return campaign_analysis